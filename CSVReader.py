import openpyxl
import json

filename = "IR1_7k_news.xlsx"


def getCSVData():
    dataDict = {}
    titlesDict = {}

    wookbook = openpyxl.load_workbook(filename)

    worksheet = wookbook.active

    for i in range(1, worksheet.max_row):

        for col in worksheet.iter_cols(0, 1):
            dataDict[i - 1] = col[i].value
        for col in worksheet.iter_cols(3):
            titlesDict[i - 1] = col[i].value

    with open('dataDict', 'w', encoding='utf-8') as convert_file:
        convert_file.write(json.dumps(dataDict))

    with open('titlesDict', 'w', encoding='utf-8') as convert_file:
        convert_file.write(json.dumps(titlesDict))


getCSVData()
