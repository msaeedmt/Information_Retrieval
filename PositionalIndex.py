import math

from hazm import *
from PostingList import PostingList
from Utils import *
import os.path
import json
import openpyxl
import pickle
import matplotlib.pyplot as plt
from Posting import Posting
import heapq
from gensim.models import Word2Vec
import random


class PositionalIndex:
    def __init__(self):
        self.xlsxFileName = "IR1_7k_news.xlsx"

        self.normalizer = Normalizer()
        self.stemmer = Stemmer()

        self.dataDict = {}
        self.titlesDict = {}
        self.positionalIndexDict = {}

        self.docsTokensList = []
        self.docs_tf_idf = []
        self.docs_embedding = []

        self.word2vector = None

        self.numberOfTokens = 0
        self.numberOfDocs = 0

        self.getXlsxData('all')
        self.getPositionalIndex()

    def getXlsxData(self, rowCounts):
        if os.path.exists("dataDict") and os.path.exists("titlesDict"):
            with open('titlesDict', encoding='utf-8') as json_file:
                self.titlesDict = json.load(json_file)
            with open('dataDict', encoding='utf-8') as json_file:
                self.dataDict = json.load(json_file)
        else:
            workbook = openpyxl.load_workbook(self.xlsxFileName)
            worksheet = workbook.active

            if rowCounts == 'all':
                rowCounts = worksheet.max_row
            for i in range(1, rowCounts):
                for col in worksheet.iter_cols(0, 1):
                    self.dataDict[i - 1] = col[i].value
                for col in worksheet.iter_cols(3):
                    self.titlesDict[i - 1] = col[i].value

            with open('dataDict', 'w', encoding='utf-8') as convert_file:
                convert_file.write(json.dumps(self.dataDict))
            with open('titlesDict', 'w', encoding='utf-8') as convert_file:
                convert_file.write(json.dumps(self.titlesDict))

        self.numberOfDocs = len(self.dataDict)

    def getPositionalIndex(self):
        if os.path.exists("PostingList.dat"):
            inputFile = open('PostingList.dat', 'rb')
            self.positionalIndexDict = pickle.load(inputFile)
            inputFile.close()
        else:
            for doc in self.dataDict:
                self.dataDict[doc] = self.normalizer.normalize(self.dataDict[doc])
                docWordsList = word_tokenize(self.dataDict[doc])

                for stopWord in stopwords_list():
                    docWordsList = list(filter(stopWord.__ne__, docWordsList))

                for i in range(len(docWordsList)):
                    docWordsList[i] = self.stemmer.stem(docWordsList[i])
                    self.numberOfTokens += 1

                    if docWordsList[i] in self.positionalIndexDict:
                        if doc in self.positionalIndexDict[docWordsList[i]].postingsList:
                            self.positionalIndexDict[docWordsList[i]].postingsList[doc].occurrences.append(int(i + 1))
                        else:
                            # newPostingList = [int(i + 1)]
                            self.positionalIndexDict[docWordsList[i]].postingsList[doc] = Posting()
                            self.positionalIndexDict[docWordsList[i]].postingsList[doc].occurrences.append(int(i + 1))
                            self.positionalIndexDict[docWordsList[i]].numberOfDocs += 1
                    else:
                        self.positionalIndexDict[docWordsList[i]] = PostingList()

                        # newPostingList = [int(i + 1)]
                        self.positionalIndexDict[docWordsList[i]].postingsList[doc] = Posting()
                        self.positionalIndexDict[docWordsList[i]].postingsList[doc].occurrences.append(int(i + 1))
                        self.positionalIndexDict[docWordsList[i]].numberOfDocs += 1

                    self.positionalIndexDict[docWordsList[i]].postingsList[doc].occurrencesCount += 1

            self.makeChampionsList(10)
            print('Serializing PositionalIndexDict ...')

            outputFile = open('PostingList.dat', 'wb')
            pickle.dump(self.positionalIndexDict, outputFile)
            outputFile.close()

            print('PositionalIndexDict has gotten successfully!')

    def makeChampionsList(self, r):
        for word in self.positionalIndexDict:
            wordDict = self.positionalIndexDict[word].postingsList
            self.positionalIndexDict[word].championList = heapq.nlargest(r, wordDict,
                                                                         key=lambda x: wordDict[x].occurrencesCount)

    def startGettingQueries(self):
        while True:
            userQuery = input("Please enter your query > ")
            userQuery = self.normalizer.normalize(userQuery)
            splitedUserQuery = userQuery.split()
            splitedUserQueriesWithoutStopWords = []
            for word in splitedUserQuery:
                if word not in stopwords_list():
                    splitedUserQueriesWithoutStopWords.append(word)

            userQueryWordsList = list(map(lambda item: self.stemmer.stem(item), splitedUserQueriesWithoutStopWords))

            print(userQueryWordsList)

            if len(userQueryWordsList) == 0:
                print('All documents')
            elif len(userQueryWordsList) == 1:
                if userQueryWordsList[0] in self.positionalIndexDict:
                    print(self.positionalIndexDict[userQueryWordsList[0]].postingsList.keys())

                    releventDocumentsCount = 0
                    for docId in self.positionalIndexDict[userQueryWordsList[0]].postingsList.keys():
                        if releventDocumentsCount == 10:
                            break
                        else:
                            releventDocumentsCount += 1

                        docId = str(docId)
                        documentContent = self.dataDict[docId]
                        realTokenizedSentences = sent_tokenize(documentContent)
                        normalizedTokenizedSentences = []

                        for j in range(len(realTokenizedSentences)):
                            normalizedSentence = self.normalizer.normalize(realTokenizedSentences[j])
                            splitedSentence = word_tokenize(normalizedSentence)
                            for stopWord in stopwords_list():
                                splitedSentence = list(filter(stopWord.__ne__, splitedSentence))
                            for t in range(len(splitedSentence)):
                                splitedSentence[t] = self.stemmer.stem(splitedSentence[t])
                            normalizedTokenizedSentences.append(' '.join(splitedSentence))

                        foundedSentencesIndexes = set()

                        for t in range(len(normalizedTokenizedSentences)):
                            if userQueryWordsList[0] in normalizedTokenizedSentences[t]:
                                foundedSentencesIndexes.add(t)

                        print('شماره سطر اکسل : ', docId)
                        print('تیتر خبر : ', self.titlesDict[docId])
                        print('جمله(جملات) : ')
                        for index in foundedSentencesIndexes:
                            print(realTokenizedSentences[index])

                        print('\n')
                else:
                    print("No result found!")
            else:
                allReleventResultsWithOrderedSubQueries = []

                fullQueryPostingList = {}
                for word in userQueryWordsList:
                    if word in self.positionalIndexDict:
                        fullQueryPostingList[word] = self.positionalIndexDict[word]

                allWordExistenceDocuments, releventResultsWithOrderedSubQueries = findQuery(fullQueryPostingList,
                                                                                            userQueryWordsList)
                allReleventResultsWithOrderedSubQueries.append(releventResultsWithOrderedSubQueries)

                for i in range(1, len(userQueryWordsList) - 1):
                    releventResultsWithOrderedSubQueries = set()
                    for j in range(i + 1):
                        print(userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                        currentSubQueryReleventdocs = \
                            findQuery(fullQueryPostingList,
                                      userQueryWordsList[j:(len(userQueryWordsList) - i + j + 1)])[1]
                        print(currentSubQueryReleventdocs)
                        releventResultsWithOrderedSubQueries.update(currentSubQueryReleventdocs)

                    for releventSubQuerySet in allReleventResultsWithOrderedSubQueries:
                        releventResultsWithOrderedSubQueries = releventResultsWithOrderedSubQueries - releventSubQuerySet

                    allReleventResultsWithOrderedSubQueries.append(releventResultsWithOrderedSubQueries)

                print()
                print('allReleventResultsWithOrderedSubQueries : ', allReleventResultsWithOrderedSubQueries)
                print('allWordExistenceDocuments : ', allWordExistenceDocuments)
                allReleventResultsWithOrderedSubQueries.append(allWordExistenceDocuments)

                releventDocumentsCount = 0
                for i in range(len(allReleventResultsWithOrderedSubQueries)):
                    if i == len(allReleventResultsWithOrderedSubQueries) - 1:
                        print('-------------------- Without Order ---------------------')
                    if releventDocumentsCount == 10:
                        break
                    for docId in allReleventResultsWithOrderedSubQueries[i]:
                        if releventDocumentsCount == 10:
                            break
                        else:
                            releventDocumentsCount += 1

                        docId = str(docId)
                        documentContent = self.dataDict[docId]
                        realTokenizedSentences = sent_tokenize(documentContent)
                        normalizedTokenizedSentences = []

                        for j in range(len(realTokenizedSentences)):
                            normalizedSentence = self.normalizer.normalize(realTokenizedSentences[j])
                            splitedSentence = word_tokenize(normalizedSentence)
                            for stopWord in stopwords_list():
                                splitedSentence = list(filter(stopWord.__ne__, splitedSentence))
                            for t in range(len(splitedSentence)):
                                splitedSentence[t] = self.stemmer.stem(splitedSentence[t])
                            normalizedTokenizedSentences.append(' '.join(splitedSentence))

                        foundedSentencesIndexes = set()

                        for j in range(i + 1):
                            print('\n', userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                            subQuery = ' '.join(userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                            for t in range(len(normalizedTokenizedSentences)):
                                if subQuery in normalizedTokenizedSentences[t]:
                                    foundedSentencesIndexes.add(t)

                        print('شماره سطر اکسل : ', docId)
                        print('تیتر خبر : ', self.titlesDict[docId])
                        print('جمله(جملات) : ')
                        for index in foundedSentencesIndexes:
                            print(realTokenizedSentences[index])

                        print('\n')

    def resolveQueriesWithVectorSpace(self):
        while True:
            userQuery = input("Please enter your query > ")
            userQuery = self.normalizer.normalize(userQuery)
            splitedUserQuery = userQuery.split()
            splitedUserQueriesWithoutStopWords = []
            for word in splitedUserQuery:
                if word not in stopwords_list():
                    splitedUserQueriesWithoutStopWords.append(word)

            userQueryWordsList = list(map(lambda item: self.stemmer.stem(item), splitedUserQueriesWithoutStopWords))

            docScores = [0] * self.numberOfDocs
            docSquareLengths = [0] * self.numberOfDocs

            userQueryWordsDict = {}
            for word in userQueryWordsList:
                if word in userQueryWordsDict:
                    userQueryWordsDict[word] = userQueryWordsDict[word] + 1
                else:
                    userQueryWordsDict[word] = 1
            for queryWord in userQueryWordsDict:
                tf_tq = userQueryWordsDict[queryWord]
                df_qt = self.positionalIndexDict[queryWord].numberOfDocs
                if 3 > (self.numberOfDocs / df_qt):
                    continue
                w_tq = calcLTC(tf_tq, df_qt, self.numberOfDocs)
                for docId in self.positionalIndexDict[queryWord].championList:
                    w_td = calcLNC(self.positionalIndexDict[queryWord].postingsList[docId].occurrencesCount)
                    docScores[int(docId)] += w_td * w_tq
                    docSquareLengths[int(docId)] += w_td ** 2
            docLengths = np.sqrt(docSquareLengths)
            for i in range(len(docScores)):
                if docScores[i] != 0:
                    docScores[i] = docScores[i] / docLengths[i]

            relatedDocs = FirstKelements(docScores, 10)
            print(relatedDocs)

            for docId in relatedDocs:
                docId = str(docId)
                documentContent = self.dataDict[docId]
                realTokenizedSentences = sent_tokenize(documentContent)
                normalizedTokenizedSentences = []

                for j in range(len(realTokenizedSentences)):
                    normalizedSentence = self.normalizer.normalize(realTokenizedSentences[j])
                    splitedSentence = word_tokenize(normalizedSentence)
                    for stopWord in stopwords_list():
                        splitedSentence = list(filter(stopWord.__ne__, splitedSentence))
                    for t in range(len(splitedSentence)):
                        splitedSentence[t] = self.stemmer.stem(splitedSentence[t])
                    normalizedTokenizedSentences.append(' '.join(splitedSentence))

                foundedSentencesIndexes = set()

                for i in range(len(userQueryWordsList)):
                    for j in range(i + 1):
                        # print('\n', userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                        subQuery = ' '.join(userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                        for t in range(len(normalizedTokenizedSentences)):
                            if subQuery in normalizedTokenizedSentences[t]:
                                foundedSentencesIndexes.add(t)

                print('شماره سطر اکسل : ', docId)
                print('تیتر خبر : ', self.titlesDict[docId])
                print('جمله(جملات) : ')
                for index in foundedSentencesIndexes:
                    print(realTokenizedSentences[index])

                print('\n')

    def resolveQueryWithWordEmbedding(self):
        if os.path.exists("docsTokenizedList.dat"):
            docsTokenizedListFile = open('docsTokenizedList.dat', 'rb')
            self.docsTokensList = pickle.load(docsTokenizedListFile)
            docsTokenizedListFile.close()
        else:
            for doc in self.dataDict:
                self.dataDict[doc] = self.normalizer.normalize(self.dataDict[doc])
                docWordsList = word_tokenize(self.dataDict[doc])

                for stopWord in stopwords_list():
                    docWordsList = list(filter(stopWord.__ne__, docWordsList))

                for i in range(len(docWordsList)):
                    docWordsList[i] = self.stemmer.stem(docWordsList[i])

                self.docsTokensList.append(docWordsList)

                print('Serializing docsTokenizedList ...')

                outputFile = open('docsTokenizedList.dat', 'wb')
                pickle.dump(self.docsTokensList, outputFile)
                outputFile.close()

                print('docsTokenizedList has gotten successfully!')

        # choos the model to train with ***************************
        self.trainData()
         # ********************************************************

        if os.path.exists("docs_tf_idf.dat"):
            docs_tf_idf_file = open("docs_tf_idf.dat", 'rb')
            self.docs_tf_idf = pickle.load(docs_tf_idf_file)
            docs_tf_idf_file.close()
        else:
            for docId in range(len(self.docsTokensList)):
                doc_tf_idf = {}
                for word in self.docsTokensList[docId]:
                    w_td = calcLTC(self.positionalIndexDict[word].postingsList[str(docId)].occurrencesCount,
                                   self.positionalIndexDict[word].numberOfDocs, self.numberOfDocs)
                    doc_tf_idf[word] = w_td
                self.docs_tf_idf.append(doc_tf_idf)

            print('Serializing docs_tf_idf ...')

            outputFile = open('docs_tf_idf.dat', 'wb')
            pickle.dump(self.docs_tf_idf, outputFile)
            outputFile.close()

            print('docs_tf_idf has gotten successfully!')

        for doc in self.docs_tf_idf:
            docVec = np.zeros(300)
            weightsSum = 0
            for token, weight in doc.items():
                if token not in self.word2vec.wv:
                    self.word2vec.wv[token] = np.ones(300) * 0.01
                docVec += self.word2vec.wv[token] * weight
                weightsSum += weight
            self.docs_embedding.append(docVec / weightsSum)

        # getting user queries
        while True:
            userQuery = input("Please enter your query > ")
            userQuery = self.normalizer.normalize(userQuery)
            splitedUserQuery = userQuery.split()
            splitedUserQueriesWithoutStopWords = []
            for word in splitedUserQuery:
                if word not in stopwords_list():
                    splitedUserQueriesWithoutStopWords.append(word)

            userQueryWordsList = list(map(lambda item: self.stemmer.stem(item), splitedUserQueriesWithoutStopWords))

            userQueryWordsDict = {}
            queryWords_df_itf = {}
            for word in userQueryWordsList:
                if word in userQueryWordsDict:
                    userQueryWordsDict[word] = userQueryWordsDict[word] + 1
                else:
                    userQueryWordsDict[word] = 1
            for queryWord in userQueryWordsDict:
                tf_tq = userQueryWordsDict[queryWord]
                df_qt = self.positionalIndexDict[queryWord].numberOfDocs
                w_tq = calcLTC(tf_tq, df_qt, self.numberOfDocs)
                queryWords_df_itf[queryWord] = w_tq

            queryDocVec = np.zeros(300)
            queryWeightsSum = 0
            for token, weight in queryWords_df_itf.items():
                if token not in self.word2vec.wv:
                    self.word2vec.wv[token] = np.ones(300) * 0.01
                queryDocVec += self.word2vec.wv[token] * weight
                queryWeightsSum += weight
            queryEmbedding = queryDocVec / queryWeightsSum

            docScores = []
            for doc in self.docs_embedding:
                docScores.append(docsSimilarity(queryEmbedding, doc))
            relatedDocs = FirstKelements(docScores, 10)
            for index in relatedDocs:
                print(docScores[index], end='--')

            print(relatedDocs)

            for docId in relatedDocs:
                docId = str(docId)
                documentContent = self.dataDict[docId]
                realTokenizedSentences = sent_tokenize(documentContent)
                normalizedTokenizedSentences = []

                for j in range(len(realTokenizedSentences)):
                    normalizedSentence = self.normalizer.normalize(realTokenizedSentences[j])
                    splitedSentence = word_tokenize(normalizedSentence)
                    for stopWord in stopwords_list():
                        splitedSentence = list(filter(stopWord.__ne__, splitedSentence))
                    for t in range(len(splitedSentence)):
                        splitedSentence[t] = self.stemmer.stem(splitedSentence[t])
                    normalizedTokenizedSentences.append(' '.join(splitedSentence))

                foundedSentencesIndexes = set()

                for i in range(len(userQueryWordsList)):
                    for j in range(i + 1):
                        # print('\n', userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                        subQuery = ' '.join(userQueryWordsList[j:(len(userQueryWordsList) - i + j)])
                        for t in range(len(normalizedTokenizedSentences)):
                            if subQuery in normalizedTokenizedSentences[t]:
                                foundedSentencesIndexes.add(t)

                print('شماره سطر اکسل : ', docId)
                print('تیتر خبر : ', self.titlesDict[docId])
                print('جمله(جملات) : ')
                for index in foundedSentencesIndexes:
                    print(realTokenizedSentences[index])

                print('\n')

    def trainData(self, model="xlsx"):
        # print('number all docs : ', len(self.docsTokensList))
        # print('number all tokens : ', sum([len(x) for x in self.docsTokensList]))
        if model == "xlsx":
            if os.path.exists("word2vector.model"):
                self.word2vec = Word2Vec.load('word2vector.model')
            else:
                self.word2vec = Word2Vec(min_count=1, window=5, vector_size=300, alpha=0.03, workers=7)
                self.word2vec.build_vocab(self.docsTokensList)
                self.word2vec.train(self.docsTokensList, total_examples=self.word2vec.corpus_count, epochs=25)
                self.word2vec.save("word2vector.model")
        elif model == "hazm":
            self.word2vec = Word2Vec.load('word2vec_model_hazm/w2v_150k_hazm_300_v2.model')

    def kmeans(self,k, datapoints):

        # d - Dimensionality of Datapoints
        d = len(datapoints[0])

        # Limit our iterations
        Max_Iterations = 1000
        i = 0

        cluster = [0] * len(datapoints)
        prev_cluster = [-1] * len(datapoints)

        # Randomly Choose Centers for the Clusters
        cluster_centers = []
        for i in range(0, k):
            new_cluster = []
            # for i in range(0,d):
            #    new_cluster += [random.randint(0,10)]
            cluster_centers += [random.choice(datapoints)]

            # Sometimes The Random points are chosen poorly and so there ends up being empty clusters
            # In this particular implementation we want to force K exact clusters.
            # To take this feature off, simply take away "force_recalculation" from the while conditional.
            force_recalculation = False

        while (cluster != prev_cluster) or (i > Max_Iterations) or (force_recalculation):

            prev_cluster = list(cluster)
            force_recalculation = False
            i += 1

            # Update Point's Cluster Alligiance
            for p in range(0, len(datapoints)):
                min_dist = float("inf")

                # Check min_distance against all centers
                for c in range(0, len(cluster_centers)):

                    dist = eucldist(datapoints[p], cluster_centers[c])

                    if (dist < min_dist):
                        min_dist = dist
                        cluster[p] = c  # Reassign Point to new Cluster

            # Update Cluster's Position
            for k in range(0, len(cluster_centers)):
                new_center = [0] * d
                members = 0
                for p in range(0, len(datapoints)):
                    if (cluster[p] == k):  # If this point belongs to the cluster
                        for j in range(0, d):
                            new_center[j] += datapoints[p][j]
                        members += 1

                for j in range(0, d):
                    if members != 0:
                        new_center[j] = new_center[j] / float(members)

                        # This means that our initial random assignment was poorly chosen
                    # Change it to a new datapoint to actually force k clusters
                    else:
                        new_center = random.choice(datapoints)
                        force_recalculation = True
                        print
                        "Forced Recalculation..."

                cluster_centers[k] = new_center

        print
        "======== Results ========"
        print
        "Clusters", cluster_centers
        print
        "Iterations", i
        print
        "Assignments", cluster

    def plotZipfs(self):
        cfisList = list(
            map(lambda item: math.log(countWordFreq(self.positionalIndexDict[item].postingsList), 10),
                self.positionalIndexDict))
        x = list(map(lambda item: math.log(item, 10), list(range(1, len(self.positionalIndexDict) + 1))))
        mergeSort(cfisList)

        logK = cfisList[0]
        estimatedList = list(
            map(lambda item: logK - math.log(item, 10), list(range(1, len(self.positionalIndexDict) + 1))))

        plt.plot(x, cfisList, color='r')
        plt.plot(x, estimatedList, color='g')

        plt.xlabel('log10 rank')
        plt.ylabel('log10 cf')

        plt.show()
